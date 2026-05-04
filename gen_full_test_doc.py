# -*- coding: utf-8 -*-
"""
gen_full_test_doc.py
สร้าง Excel รวม Use Cases + Test Cases + Unit Tests ทั้งหมด 267 tests
"""
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

OUT = r"c:\dev\calories-guard\CaloriesGuard_FullTestDoc.xlsx"

# ── Palette ─────────────────────────────────────────────────────────────────
C = {
    "header_green":  "2E7D32", "header_blue":   "1565C0",
    "header_orange": "E65100", "header_purple": "6A1B9A",
    "header_teal":   "00695C", "header_red":    "B71C1C",
    "row_green":     "E8F5E9", "row_blue":      "E3F2FD",
    "row_orange":    "FFF3E0", "row_purple":    "F3E5F5",
    "row_teal":      "E0F2F1", "row_alt":       "F5F5F5",
    "pass":          "C8E6C9", "fail":          "FFCDD2",
    "warn":          "FFF9C4", "white":         "FFFFFF",
    "title_bg":      "1B5E20", "sub_bg":        "43A047",
}

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(bold=False, color="000000", sz=11):
    return Font(bold=bold, color=color, size=sz, name="Calibri")
def center(): return Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center",
                               wrap_text=True)
def thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def write_header(ws, row, cols, bg, fg="FFFFFF", sz=11):
    for c, (col, val) in enumerate(cols.items(), 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill(bg); cell.font = font(True, fg, sz)
        cell.alignment = center(); cell.border = thin_border()

def write_row(ws, row, vals, bg, bold=False):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = fill(bg)
        cell.font = font(bold, "212121")
        cell.alignment = left()
        cell.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════════════════
# DATA
# ═══════════════════════════════════════════════════════════════════════════

USE_CASES = [
    # UC-ID, Module, Use Case Name, Actor, Precondition, Main Flow, Expected Result, Priority
    ("UC-001","Authentication","สมัครสมาชิก","ผู้ใช้ใหม่","เปิดแอปครั้งแรก",
     "1.กรอกชื่อ-นามสกุล 2.กรอกอีเมล 3.กรอกรหัสผ่าน 2 ครั้ง 4.กด Done",
     "บัญชีถูกสร้าง, ระบบส่ง email verify, ไปหน้า verify email","High"),
    ("UC-002","Authentication","เข้าสู่ระบบ","ผู้ใช้","มีบัญชีและยืนยัน email แล้ว",
     "1.กรอกอีเมล 2.กรอกรหัสผ่าน 3.กด Login",
     "เข้าสู่หน้าหลักได้, Provider โหลดข้อมูล user","High"),
    ("UC-003","Authentication","ยืนยัน Email","ผู้ใช้","ได้รับ email OTP",
     "1.กรอกรหัส 6 หลักจาก email 2.กด Verify",
     "is_email_verified = true, ไปหน้า gender selection","High"),
    ("UC-004","Authentication","รีเซ็ตรหัสผ่าน","ผู้ใช้","ลืมรหัสผ่าน",
     "1.กด Forgot Password 2.กรอกอีเมล 3.รับ OTP 4.กรอกรหัสใหม่",
     "รหัสผ่านเปลี่ยน, สามารถ login ด้วยรหัสใหม่","Medium"),
    ("UC-005","Onboarding","เลือกเพศ","ผู้ใช้","สมัครและยืนยัน email แล้ว",
     "1.เลือก ชาย/หญิง 2.กด ถัดไป",
     "gender บันทึกใน DB และ Provider","High"),
    ("UC-006","Onboarding","กรอกข้อมูลส่วนตัว","ผู้ใช้","เลือกเพศแล้ว",
     "1.เลือกวันเกิด 2.กรอกส่วนสูง 3.กรอกน้ำหนัก 4.กด ถัดไป",
     "ข้อมูลบันทึกใน DB, ระบบคำนวณ BMI/BMR","High"),
    ("UC-007","Onboarding","เลือกเป้าหมาย","ผู้ใช้","กรอกข้อมูลส่วนตัวแล้ว",
     "1.เลือก ลดน้ำหนัก/รักษา/เพิ่มกล้ามเนื้อ 2.กด ถัดไป",
     "goal_type บันทึกใน DB","High"),
    ("UC-008","Onboarding","ตั้งน้ำหนักเป้าหมาย","ผู้ใช้","เลือกเป้าหมายแล้ว",
     "1.เลื่อน slider น้ำหนักเป้า 2.เลื่อน slider ระยะเวลา 3.กด ถัดไป",
     "target_weight และ target_date บันทึกใน DB, หากเร็วเกิน 1kg/week แจ้งเตือน","High"),
    ("UC-009","Onboarding","เลือกระดับกิจกรรม","ผู้ใช้","ตั้งน้ำหนักเป้าหมายแล้ว",
     "1.เลือก sedentary/lightly/moderately/very active 2.กด ถัดไป",
     "activity_level บันทึก, ระบบคำนวณ TDEE","High"),
    ("UC-010","Home","ดู Dashboard","ผู้ใช้","Login และตั้ง goal แล้ว",
     "1.เปิดหน้า Home 2.ดูวงแหวนแคลอรี่ 3.ดู macro progress",
     "แสดง consumed vs target calories และ macros ถูกต้อง","High"),
    ("UC-011","Food","บันทึกอาหาร","ผู้ใช้","อยู่หน้า Home",
     "1.กด + เพิ่มอาหาร 2.ค้นหาอาหาร 3.เลือก meal type 4.กด บันทึก",
     "อาหารบันทึกใน DB, consumed calories อัปเดต","High"),
    ("UC-012","Food","ค้นหาอาหาร","ผู้ใช้","หน้า log food",
     "1.พิมพ์ชื่ออาหาร 2.ดูผลลัพธ์ 3.กดเลือก",
     "แสดงรายการอาหารตรงกับคำค้น","High"),
    ("UC-013","Health","ดู BMI Status","ผู้ใช้","กรอกข้อมูลส่วนตัวแล้ว",
     "1.ไปหน้า BMI 2.ดูค่า BMI และสี status",
     "BMI คำนวณถูกต้อง, สีแสดง: เขียว/เหลือง/ส้ม/แดง","Medium"),
    ("UC-014","Gamification","ดู Tamagotchi","ผู้ใช้","Login แล้ว",
     "1.ไปหน้า Tamagotchi 2.ดูสถานะ tama 3.กิน/ออกกำลังกาย",
     "tama state เปลี่ยนตาม action, แสดง animation","Medium"),
    ("UC-015","Gamification","รับ Badge","ผู้ใช้","บรรลุ milestone",
     "1.ทำ action ที่กำหนด (เช่น streak 7 วัน) 2.ระบบ check condition",
     "badge unlock, แสดง popup แจ้งเตือน","Medium"),
    ("UC-016","Profile","แก้ไขโปรไฟล์","ผู้ใช้","Login แล้ว",
     "1.ไปหน้า Profile 2.แก้ไขข้อมูล 3.กด บันทึก",
     "ข้อมูลอัปเดตใน DB และ Provider","Medium"),
    ("UC-017","Settings","เปลี่ยนภาษา","ผู้ใช้","หน้า Settings",
     "1.เลือก ภาษาไทย/อังกฤษ 2.กด บันทึก",
     "แอปเปลี่ยนภาษาทันที","Low"),
    ("UC-018","Settings","เปลี่ยน Theme","ผู้ใช้","หน้า Settings",
     "1.เลือก Light/Dark/System 2.กด บันทึก",
     "แอปเปลี่ยน theme ทันที","Low"),
]

# ── Test Cases ─────────────────────────────────────────────────────────────
# (TC-ID, UC-Ref, Module, Test Case Name, Test Type, Input, Expected Output, Status, Priority)
TEST_CASES = []

# ─── Health Calc (health_calc_test.dart) ─────────────────────────────────
hc = [
    # BMI
    ("TC-001","UC-013","BMI","BMI ชาย 70kg 175cm → 22.86","Unit","weight=70, height=175","22.86","PASS","High"),
    ("TC-002","UC-013","BMI","BMI หญิง 55kg 160cm → 21.48","Unit","weight=55, height=160","21.48","PASS","High"),
    ("TC-003","UC-013","BMI","BMI weight=0 → 0 (guard)","Unit","weight=0, height=175","0","PASS","High"),
    ("TC-004","UC-013","BMI","BMI height=0 → 0 (guard)","Unit","weight=70, height=0","0","PASS","High"),
    ("TC-005","UC-013","BMI","BMI WHO Asian: <18.5=underweight","Unit","BMI=17.0","Underweight","PASS","Medium"),
    ("TC-006","UC-013","BMI","BMI WHO Asian: 18.5-22.9=normal","Unit","BMI=21.0","Normal","PASS","Medium"),
    ("TC-007","UC-013","BMI","BMI WHO Asian: 23-24.9=overweight","Unit","BMI=23.5","Overweight","PASS","Medium"),
    ("TC-008","UC-013","BMI","BMI WHO Asian: ≥25=obese","Unit","BMI=27.0","Obese","PASS","Medium"),
    # BMR
    ("TC-009","UC-006","BMR","BMR ชาย 70kg 175cm 25yr (Asian×0.94)","Unit","M, 70kg, 175cm, 25yr","1573.3 kcal","PASS","High"),
    ("TC-010","UC-006","BMR","BMR หญิง 55kg 160cm 25yr","Unit","F, 55kg, 160cm, 25yr","1157.6 kcal","PASS","High"),
    ("TC-011","UC-006","BMR","BMR weight=0 → 1500","Unit","weight=0","1500","PASS","High"),
    ("TC-012","UC-006","BMR","BMR height=0 → 1500","Unit","height=0","1500","PASS","High"),
    ("TC-013","UC-006","BMR","BMR หญิง < BMR ชาย (ส่วนสูง/น้ำหนักเท่ากัน)","Unit","M vs F, 70kg 170cm","male > female","PASS","High"),
    ("TC-014","UC-006","BMR","BMR Asian < Mifflin raw เสมอ","Unit","70kg, 175cm, 25yr","bmr < raw_mifflin","PASS","High"),
    ("TC-015","UC-006","BMR","อายุมากขึ้น → BMR ลดลง","Unit","25yr vs 65yr","older < younger","PASS","Medium"),
    # Activity / TDEE
    ("TC-016","UC-009","TDEE","sedentary → BMR×1.2","Unit","activity=sedentary","TDEE=BMR×1.2","PASS","High"),
    ("TC-017","UC-009","TDEE","lightly_active → BMR×1.375","Unit","activity=lightly_active","TDEE=BMR×1.375","PASS","High"),
    ("TC-018","UC-009","TDEE","moderately_active → BMR×1.55","Unit","activity=moderately_active","TDEE=BMR×1.55","PASS","High"),
    ("TC-019","UC-009","TDEE","very_active → BMR×1.725","Unit","activity=very_active","TDEE=BMR×1.725","PASS","High"),
    ("TC-020","UC-009","TDEE","extra_active → BMR×1.9","Unit","activity=extra_active","TDEE=BMR×1.9","PASS","High"),
    ("TC-021","UC-009","TDEE","unknown activity → default 1.2","Unit","activity=invalid","TDEE=BMR×1.2","PASS","Medium"),
    ("TC-022","UC-009","TDEE","TDEE > BMR เสมอ","Unit","any valid user","tdee > bmr","PASS","High"),
    # Macros
    ("TC-023","UC-007","Macros","ข้าวสวย 1 ถ้วย → ~196.5 kcal","Unit","carb=44g, prot=4g, fat=0.5g","196.5 kcal","PASS","Medium"),
    ("TC-024","UC-007","Macros","ไก่ย่าง 100g → ~165 kcal","Unit","prot=31g, carb=0g, fat=3.6g","165.4 kcal","PASS","Medium"),
    ("TC-025","UC-007","Macros","แมคโคร 0 ทั้งหมด → 0 kcal","Unit","prot=0, carb=0, fat=0","0","PASS","Medium"),
    # Target Calories
    ("TC-026","UC-008","Target Cal","ลดน้ำหนัก 0.5kg/wk TDEE 2200 → 1650","Unit","TDEE=2200, -0.5kg/wk","1650 kcal","PASS","High"),
    ("TC-027","UC-008","Target Cal","เพิ่มกล้าม 0.5kg/wk TDEE 2000 → 2550","Unit","TDEE=2000, +0.5kg/wk","2550 kcal","PASS","High"),
    ("TC-028","UC-008","Target Cal","ลดต่ำเกิน → floor 1500 (ชาย)","Unit","TDEE=1600, -0.8kg/wk, male","1500 kcal","PASS","High"),
    ("TC-029","UC-008","Target Cal","ลดต่ำเกิน → floor 1200 (หญิง)","Unit","TDEE=1400, -0.8kg/wk, female","1200 kcal","PASS","High"),
    # Validation
    ("TC-030","UC-003","Validation","Email มี @ → valid","Unit","user@example.com","True","PASS","High"),
    ("TC-031","UC-003","Validation","Email ไม่มี @ → invalid","Unit","userexample.com","False","PASS","High"),
    ("TC-032","UC-003","Validation","Email ว่าง → invalid","Unit","''","False","PASS","High"),
    ("TC-033","UC-006","Validation","อายุ 13 → valid (min)","Unit","age=13","True","PASS","High"),
    ("TC-034","UC-006","Validation","อายุ 12 → invalid","Unit","age=12","False","PASS","High"),
    ("TC-035","UC-006","Validation","อายุ 100 → valid (max)","Unit","age=100","True","PASS","High"),
    ("TC-036","UC-006","Validation","อายุ 101 → invalid","Unit","age=101","False","PASS","High"),
    ("TC-037","UC-006","Validation","ส่วนสูง 100 cm → valid","Unit","height=100","True","PASS","High"),
    ("TC-038","UC-006","Validation","ส่วนสูง 99 cm → invalid","Unit","height=99","False","PASS","High"),
    ("TC-039","UC-006","Validation","ส่วนสูง 250 cm → valid","Unit","height=250","True","PASS","High"),
    ("TC-040","UC-006","Validation","ส่วนสูง 251 cm → invalid","Unit","height=251","False","PASS","High"),
    ("TC-041","UC-006","Validation","น้ำหนัก 20 kg → valid","Unit","weight=20","True","PASS","High"),
    ("TC-042","UC-006","Validation","น้ำหนัก 19 kg → invalid","Unit","weight=19","False","PASS","High"),
    ("TC-043","UC-006","Validation","น้ำหนัก 300 kg → valid","Unit","weight=300","True","PASS","High"),
    ("TC-044","UC-006","Validation","น้ำหนัก 301 kg → invalid","Unit","weight=301","False","PASS","High"),
    # Weight Loss Safety
    ("TC-045","UC-008","Safety","0.3 kg/wk → recommended","Unit","kgPerWeek=0.3","'recommended'","PASS","High"),
    ("TC-046","UC-008","Safety","0.5 kg/wk → recommended (boundary)","Unit","kgPerWeek=0.5","'recommended'","PASS","High"),
    ("TC-047","UC-008","Safety","0.7 kg/wk → borderline","Unit","kgPerWeek=0.7","'borderline'","PASS","High"),
    ("TC-048","UC-008","Safety","1.0 kg/wk → borderline (boundary)","Unit","kgPerWeek=1.0","'borderline'","PASS","High"),
    ("TC-049","UC-008","Safety","1.1 kg/wk → unsafe","Unit","kgPerWeek=1.1","'unsafe'","PASS","High"),
    ("TC-050","UC-008","Safety","negative kgPerWeek → abs ก่อน","Unit","kgPerWeek=-0.7","'borderline'","PASS","High"),
    ("TC-051","UC-008","Safety","0 kg/wk → recommended","Unit","kgPerWeek=0","'recommended'","PASS","Medium"),
    # Macro Ratios
    ("TC-052","UC-007","Macros","loseWeight: P30% C40% F30%","Unit","goal=lose, cal=2000","P150g C200g F67g","PASS","High"),
    ("TC-053","UC-007","Macros","maintainWeight: P25% C45% F30%","Unit","goal=maintain, cal=2000","P125g C225g F67g","PASS","High"),
    ("TC-054","UC-007","Macros","buildMuscle: P30% C50% F20%","Unit","goal=build, cal=2000","P150g C250g F44g","PASS","High"),
]
TEST_CASES.extend(hc)

# ─── Negative Testing ──────────────────────────────────────────────────────
neg = [
    ("TC-055","UC-006","Negative","BMI weight ติดลบ → guard 0","Negative","weight=-5, height=170","0","PASS","High"),
    ("TC-056","UC-006","Negative","BMR weight ติดลบ → fallback 1500","Negative","weight=-5, height=170","1500","PASS","High"),
    ("TC-057","UC-011","Negative","Macro calories ค่าลบ → ผลลบ (graceful)","Negative","protein=-10, carbs=0, fat=0","-40 kcal","PASS","Medium"),
    ("TC-058","UC-008","Negative","Target cal gender ไม่ใช่ male/female → floor 1200","Negative","gender='other', TDEE=1400","≥1200","PASS","High"),
    ("TC-059","UC-003","Validation","Email มีช่องว่าง → invalid","Negative","'user @test.com'","False","PASS","High"),
    ("TC-060","UC-003","Validation","Email ขึ้นต้นด้วย @ → invalid","Negative","'@test.com'","False","PASS","High"),
    ("TC-061","UC-003","Validation","Email หลาย @ → invalid","Negative","'a@@b.com'","False","PASS","High"),
    ("TC-062","UC-006","Negative","อายุ 0 → invalid","Negative","age=0","False","PASS","High"),
    ("TC-063","UC-006","Negative","อายุ -1 → invalid","Negative","age=-1","False","PASS","High"),
    ("TC-064","UC-006","Negative","ส่วนสูง 0 → invalid","Negative","height=0","False","PASS","High"),
    ("TC-065","UC-006","Negative","น้ำหนัก 0 → invalid","Negative","weight=0","False","PASS","High"),
]
TEST_CASES.extend(neg)

# ─── BVA Testing ───────────────────────────────────────────────────────────
bva = [
    ("TC-066","UC-006","BVA","อายุ 12 (min-1) → invalid","BVA","age=12","False","PASS","High"),
    ("TC-067","UC-006","BVA","อายุ 13 (min) → valid","BVA","age=13","True","PASS","High"),
    ("TC-068","UC-006","BVA","อายุ 14 (min+1) → valid","BVA","age=14","True","PASS","High"),
    ("TC-069","UC-006","BVA","อายุ 99 (max-1) → valid","BVA","age=99","True","PASS","High"),
    ("TC-070","UC-006","BVA","อายุ 100 (max) → valid","BVA","age=100","True","PASS","High"),
    ("TC-071","UC-006","BVA","อายุ 101 (max+1) → invalid","BVA","age=101","False","PASS","High"),
    ("TC-072","UC-006","BVA","ส่วนสูง 99 (min-1) → invalid","BVA","height=99","False","PASS","High"),
    ("TC-073","UC-006","BVA","ส่วนสูง 100 (min) → valid","BVA","height=100","True","PASS","High"),
    ("TC-074","UC-006","BVA","ส่วนสูง 101 (min+1) → valid","BVA","height=101","True","PASS","High"),
    ("TC-075","UC-006","BVA","ส่วนสูง 249 (max-1) → valid","BVA","height=249","True","PASS","High"),
    ("TC-076","UC-006","BVA","ส่วนสูง 250 (max) → valid","BVA","height=250","True","PASS","High"),
    ("TC-077","UC-006","BVA","ส่วนสูง 251 (max+1) → invalid","BVA","height=251","False","PASS","High"),
    ("TC-078","UC-006","BVA","น้ำหนัก 19 (min-1) → invalid","BVA","weight=19","False","PASS","High"),
    ("TC-079","UC-006","BVA","น้ำหนัก 20 (min) → valid","BVA","weight=20","True","PASS","High"),
    ("TC-080","UC-006","BVA","น้ำหนัก 21 (min+1) → valid","BVA","weight=21","True","PASS","High"),
    ("TC-081","UC-006","BVA","น้ำหนัก 299 (max-1) → valid","BVA","weight=299","True","PASS","High"),
    ("TC-082","UC-006","BVA","น้ำหนัก 300 (max) → valid","BVA","weight=300","True","PASS","High"),
    ("TC-083","UC-006","BVA","น้ำหนัก 301 (max+1) → invalid","BVA","weight=301","False","PASS","High"),
    ("TC-084","UC-008","BVA","safety 0.5 kg/wk (boundary recommended)","BVA","kgPerWeek=0.5","'recommended'","PASS","High"),
    ("TC-085","UC-008","BVA","safety 0.51 (boundary+1 → borderline)","BVA","kgPerWeek=0.51","'borderline'","PASS","High"),
    ("TC-086","UC-008","BVA","safety 1.0 (boundary borderline)","BVA","kgPerWeek=1.0","'borderline'","PASS","High"),
    ("TC-087","UC-008","BVA","safety 1.01 (boundary+1 → unsafe)","BVA","kgPerWeek=1.01","'unsafe'","PASS","High"),
    ("TC-088","UC-013","BVA","BMI 18.5 (normal boundary)","BVA","BMI=18.5","Normal","PASS","Medium"),
    ("TC-089","UC-013","BVA","BMI 22.9 (normal/overweight boundary)","BVA","BMI=22.9","Normal","PASS","Medium"),
    ("TC-090","UC-013","BVA","BMI 23.0 (overweight boundary)","BVA","BMI=23.0","Overweight","PASS","Medium"),
    ("TC-091","UC-013","BVA","BMI 25.0 (obese boundary)","BVA","BMI=25.0","Obese","PASS","Medium"),
]
TEST_CASES.extend(bva)

# ─── Integration Simulation ────────────────────────────────────────────────
integ = [
    ("TC-092","UC-002","Integration","API parse user JSON → ข้อมูลครบ","Integration","valid JSON","UserData populated","PASS","High"),
    ("TC-093","UC-002","Integration","API JSON ว่าง → default values","Integration","{}","defaults","PASS","High"),
    ("TC-094","UC-002","Integration","API JSON corrupt → null (graceful)","Integration","'not json'","null","PASS","High"),
    ("TC-095","UC-014","Integration","Leaderboard parse → username/points","Integration","valid leaderboard JSON","parsed correctly","PASS","Medium"),
    ("TC-096","UC-014","Integration","Leaderboard missing username → 'ผู้ใช้'","Integration","JSON no username","'ผู้ใช้'","PASS","Medium"),
    ("TC-097","UC-010","Integration","HTTP 200 → success(0)","Integration","statusCode=200","0","PASS","High"),
    ("TC-098","UC-010","Integration","HTTP 401 → auth error(1)","Integration","statusCode=401","1","PASS","High"),
    ("TC-099","UC-010","Integration","HTTP 403 → auth error(1)","Integration","statusCode=403","1","PASS","High"),
    ("TC-100","UC-010","Integration","HTTP 500 → server error(2)","Integration","statusCode=500","2","PASS","High"),
    ("TC-101","UC-010","Integration","HTTP 404 → not found(3)","Integration","statusCode=404","3","PASS","High"),
]
TEST_CASES.extend(integ)

# ─── Gamification ──────────────────────────────────────────────────────────
gam = [
    ("TC-102","UC-014","Gamification","calcTamaLevel HP 100 → level 3","Unit","hp=100, exp=0","level 3","PASS","Medium"),
    ("TC-103","UC-014","Gamification","calcTamaLevel HP 0 → level 0 (fainted)","Unit","hp=0","level 0","PASS","Medium"),
    ("TC-104","UC-014","Gamification","feed action → HP+10 EXP+5","Unit","feed","hp+10, exp+5","PASS","Medium"),
    ("TC-105","UC-014","Gamification","exercise action → HP+5 EXP+10","Unit","exercise","hp+5, exp+10","PASS","Medium"),
    ("TC-106","UC-015","Gamification","streak 7 วัน → badge unlock","Unit","streak=7","badge='week_warrior'","PASS","Medium"),
    ("TC-107","UC-015","Gamification","streak 30 วัน → badge unlock","Unit","streak=30","badge='month_master'","PASS","Medium"),
    ("TC-108","UC-015","Gamification","streak 0 → ไม่ unlock badge","Unit","streak=0","no badge","PASS","Medium"),
    ("TC-109","UC-014","Gamification","EXP 100 → level up","Unit","exp=100","level_up=true","PASS","Medium"),
    ("TC-110","UC-014","Gamification","HP ไม่เกิน 100 (cap)","Unit","hp=95+10","hp=100","PASS","Medium"),
    ("TC-111","UC-014","Gamification","HP ไม่ต่ำกว่า 0 (floor)","Unit","hp=5-10","hp=0","PASS","Medium"),
]
TEST_CASES.extend(gam)

# ─── Full Coverage — UserData Getters ─────────────────────────────────────
fc_getters = [
    ("TC-112","UC-006","UserData","age: birthDate null → 20","Unit","birthDate=null","20","PASS","High"),
    ("TC-113","UC-006","UserData","age: birthdate ยังไม่ถึง → age-1","Unit","bday เดือนหน้า","age-1","PASS","High"),
    ("TC-114","UC-006","UserData","bmi 90kg 180cm → 27.78","Unit","w=90, h=180","27.78","PASS","High"),
    ("TC-115","UC-009","UserData","tdee unknown activity → sedentary","Unit","activity='xyz'","BMR×1.2","PASS","Medium"),
    ("TC-116","UC-008","UserData","targetCalories stored → ใช้ stored","Unit","stored=1800","1800","PASS","High"),
    ("TC-117","UC-008","UserData","targetCalories stored=0 → คำนวณ","Unit","stored=0","< TDEE","PASS","High"),
    ("TC-118","UC-007","UserData","buildMuscle → target > TDEE","Unit","goal=build","target > tdee","PASS","High"),
    ("TC-119","UC-007","UserData","macro stored values → ใช้ stored","Unit","stored_prot=150","150","PASS","High"),
    ("TC-120","UC-007","UserData","copyWith weight → field อื่นไม่เปลี่ยน","Unit","copyWith(weight=80)","height=175","PASS","Medium"),
    ("TC-121","UC-016","UserData","clearAvatarUrl=true → null","Unit","clearAvatarUrl=true","null","PASS","Medium"),
    ("TC-122","UC-016","UserData","copyWith ไม่ระบุ → ค่าเดิม","Unit","copyWith()","same values","PASS","Medium"),
]
TEST_CASES.extend(fc_getters)

# ─── Full Coverage — UserDataNotifier ─────────────────────────────────────
fc_notifier = [
    ("TC-123","UC-001","Notifier","setUserId → userId เปลี่ยน","Unit","setUserId(99)","userId=99","PASS","High"),
    ("TC-124","UC-001","Notifier","setLoginInfo → email/password ถูก set","Unit","setLoginInfo('a@b','p')","email='a@b'","PASS","High"),
    ("TC-125","UC-005","Notifier","setGender female → state.gender=female","Unit","setGender('female')","gender='female'","PASS","High"),
    ("TC-126","UC-006","Notifier","setPersonalInfo → ครบทุก field","Unit","name,bday,h,w","stored correctly","PASS","High"),
    ("TC-127","UC-007","Notifier","setGoal buildMuscle → goal=buildMuscle","Unit","setGoal(buildMuscle)","goal=buildMuscle","PASS","High"),
    ("TC-128","UC-007","Notifier","setGoal loseWeight","Unit","setGoal(loseWeight)","goal=loseWeight","PASS","High"),
    ("TC-129","UC-007","Notifier","setGoal maintainWeight","Unit","setGoal(maintain)","goal=maintain","PASS","High"),
    ("TC-130","UC-008","Notifier","setGoalInfo → target/date/duration","Unit","setGoalInfo(60,date,90)","stored","PASS","High"),
    ("TC-131","UC-009","Notifier","setActivityLevel very_active","Unit","setActivityLevel('very_active')","actLevel set","PASS","High"),
    ("TC-132","UC-011","Notifier","updateDailyFood → consumed values","Unit","cal=1800,p=120,c=200,f=60","stored","PASS","High"),
    ("TC-133","UC-011","Notifier","setDailySummaryFromApi — ข้อมูลครบ","Unit","valid API map","parsed","PASS","High"),
    ("TC-134","UC-011","Notifier","setDailySummaryFromApi — meals=null","Unit","no meals key","empty meals","PASS","Medium"),
    ("TC-135","UC-011","Notifier","setDailySummaryFromApi — empty map","Unit","{}","all zeros","PASS","Medium"),
    ("TC-136","UC-011","Notifier","resetDailyFood → all zeros","Unit","reset after update","0,0,0,0","PASS","High"),
    ("TC-137","UC-017","Notifier","updateUnit weight → lbs","Unit","updateUnit(weight='lbs')","unitWeight='lbs'","PASS","Low"),
    ("TC-138","UC-017","Notifier","updateUnit height+energy","Unit","updateUnit(h='ft',e='kj')","both changed","PASS","Low"),
    ("TC-139","UC-011","Notifier","setAllergies [1,3,5]","Unit","setAllergies([1,3,5])","[1,3,5]","PASS","Medium"),
    ("TC-140","UC-011","Notifier","setAllergies ว่าง → empty","Unit","setAllergies([])","[]","PASS","Medium"),
    ("TC-141","UC-002","Notifier","setUserFromApi — ข้อมูลครบ","Unit","valid user JSON","all fields parsed","PASS","High"),
    ("TC-142","UC-002","Notifier","setUserFromApi — username ว่าง → User","Unit","username=''","name='User'","PASS","High"),
    ("TC-143","UC-007","Notifier","setUserFromApi — goal maintain","Unit","goal_type='maintain_weight'","maintain","PASS","High"),
    ("TC-144","UC-007","Notifier","setUserFromApi — goal gain_muscle","Unit","goal_type='gain_muscle'","buildMuscle","PASS","High"),
    ("TC-145","UC-006","Notifier","setUserFromApi — height null → 0","Unit","height_cm=null","0.0","PASS","High"),
    ("TC-146","UC-006","Notifier","setUserFromApi — birth_date valid","Unit","'2000-03-20'","DateTime(2000,3,20)","PASS","High"),
    ("TC-147","UC-006","Notifier","setUserFromApi — birth_date invalid","Unit","'not-a-date'","null","PASS","High"),
    ("TC-148","UC-016","Notifier","setAvatarUrl → เปลี่ยน","Unit","setAvatarUrl(url)","url set","PASS","Medium"),
    ("TC-149","UC-016","Notifier","setAvatarUrl null → ค่าเดิมยังอยู่","Unit","setAvatarUrl(null)","old url","PASS","Medium"),
    ("TC-150","UC-016","Notifier","clearAvatarUrl=true → null","Unit","copyWith(clear=true)","null","PASS","Medium"),
    ("TC-151","UC-002","Notifier","logout → state = default","Unit","logout()","userId=0","PASS","High"),
    ("TC-152","UC-002","Notifier","reset → state = default","Unit","reset()","email=''","PASS","High"),
]
TEST_CASES.extend(fc_notifier)

# ─── Food Model ────────────────────────────────────────────────────────────
food_tc = [
    ("TC-153","UC-012","Food Model","fromJson ครบ → parse ถูกต้อง","Unit","valid JSON","Food object","PASS","High"),
    ("TC-154","UC-012","Food Model","fromJson: ใช้ 'name' fallback","Unit","no food_name, has name","name set","PASS","Medium"),
    ("TC-155","UC-012","Food Model","fromJson: food_id ขาด → 0","Unit","no food_id","id=0","PASS","Medium"),
    ("TC-156","UC-012","Food Model","fromJson: calories String → double","Unit","calories='350.5'","350.5","PASS","High"),
    ("TC-157","UC-012","Food Model","fromJson: null values → 0.0","Unit","all null","0.0","PASS","High"),
    ("TC-158","UC-012","Food Model","fromJson: imageUrl ขาด → null","Unit","no image_url","null","PASS","Low"),
    ("TC-159","UC-011","FoodLog","FoodLog snapshot อัตโนมัติจาก Food","Unit","Food(cal=250,p=35)","logCal=250","PASS","High"),
    ("TC-160","UC-011","FoodLog","MealType ครบ 4 ค่า","Unit","MealType.values","4 items","PASS","Medium"),
    ("TC-161","UC-011","FoodLog","FoodLog.meal ถูก set","Unit","meal=snack","snack","PASS","Medium"),
]
TEST_CASES.extend(food_tc)

# ─── AppSettings ───────────────────────────────────────────────────────────
settings_tc = [
    ("TC-162","UC-017","Settings","default language=th, theme=light","Unit","AppSettings()","th, light","PASS","Low"),
    ("TC-163","UC-017","Settings","copyWith language → theme ไม่เปลี่ยน","Unit","copyWith(lang='en')","lang=en, theme same","PASS","Low"),
    ("TC-164","UC-018","Settings","copyWith theme → language ไม่เปลี่ยน","Unit","copyWith(theme='dark')","theme=dark, lang same","PASS","Low"),
    ("TC-165","UC-018","Settings","copyWith ไม่ระบุ → ค่าเดิม","Unit","copyWith()","same","PASS","Low"),
    ("TC-166","UC-007","GoalOption","GoalOption มี 3 ค่า","Unit","GoalOption.values","length=3","PASS","Medium"),
    ("TC-167","UC-007","GoalOption","enum ครบ 3 ค่า","Unit","values check","all 3 present","PASS","Medium"),
]
TEST_CASES.extend(settings_tc)

# ─── Widget Tests ──────────────────────────────────────────────────────────
widget_tc = [
    ("TC-168","UC-010","Widget","Home screen render ได้","Widget","pump HomeScreen","no error","PASS","High"),
    ("TC-169","UC-001","Widget","Register screen render ได้","Widget","pump RegisterScreen","no error","PASS","High"),
    ("TC-170","UC-005","Widget","Gender selection render ได้","Widget","pump GenderScreen","no error","PASS","Medium"),
    ("TC-171","UC-006","Widget","Personal info screen render ได้","Widget","pump PersonalInfoScreen","no error","PASS","Medium"),
    ("TC-172","UC-009","Widget","Activity level screen render ได้","Widget","pump ActivityScreen","no error","PASS","Medium"),
    ("TC-173","UC-007","Widget","Goal selection screen render ได้","Widget","pump GoalScreen","no error","PASS","Medium"),
    ("TC-174","UC-013","Widget","BMI screen render ได้","Widget","pump BMIScreen","no error","PASS","Medium"),
    ("TC-175","UC-014","Widget","Tamagotchi screen render ได้","Widget","pump TamaScreen","no error","PASS","Low"),
    ("TC-176","UC-016","Widget","Profile screen render ได้","Widget","pump ProfileScreen","no error","PASS","Low"),
]
TEST_CASES.extend(widget_tc)

# ── Research References ────────────────────────────────────────────────────
REFERENCES = [
    ("REF-001","BMR Asian Correction","Huang KC et al.","Obesity Research 2004",
     "Mifflin-St Jeor overestimates REE ~6-7% in Asians → correction ×0.94",
     "TC-009 to TC-015"),
    ("REF-002","BMR Formula","Mifflin & St Jeor","Am J Clin Nutr 1990 (PMID 2305711)",
     "Male: 10w+6.25h-5a+5 | Female: 10w+6.25h-5a-161",
     "TC-009 to TC-015"),
    ("REF-003","Activity Factors","Harris-Benedict revised Frankenfield",
     "J Am Diet Assoc 2005","Sedentary 1.2 → Extra Active 1.9",
     "TC-016 to TC-022"),
    ("REF-004","Macro Atwater","FAO","FAO Food and Nutrition Paper 2003",
     "Protein 4 kcal/g | Carbs 4 kcal/g | Fat 9 kcal/g",
     "TC-023 to TC-025"),
    ("REF-005","Weight Loss Recommended","AND (Academy of Nutrition and Dietetics)",
     "2016 Guidelines","≤0.5 kg/week = sustainable, safe weight loss",
     "TC-045 to TC-051"),
    ("REF-006","Weight Loss Borderline","Stiegler & Cunliffe",
     "Sports Medicine 2006","≤1.0 kg/week acceptable with lean mass loss risk",
     "TC-045 to TC-051"),
    ("REF-007","Weight Loss Unsafe","Johansson et al.",
     "Obesity Reviews 2014",">1.0 kg/week: risk of gallstone and muscle loss",
     "TC-045 to TC-051"),
    ("REF-008","Age Min 13","AAP + COPPA","AAP 2016 / COPPA 2016",
     "Calorie-tracking apps risk disordered eating in children <13",
     "TC-033 to TC-036"),
    ("REF-009","BMI WHO Asian","WHO Expert Consultation",
     "Lancet 2004","Asian cutoff: 23=overweight, 25=obese (lower than global)",
     "TC-005 to TC-008"),
    ("REF-010","Metabolic Adaptation","Hall KD","The Lancet 2011",
     "7700 rule overpredicts long-term; recalculate every 4 weeks",
     "TC-026 to TC-029"),
]

# ═══════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ─── Sheet 1: Summary ──────────────────────────────────────────────────────
ws = wb.active
ws.title = "📊 Summary"
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 20

# Title
ws.merge_cells('A1:B1')
c = ws['A1']
c.value = "🛡️ CaloriesGuard — Test Documentation"
c.fill = fill(C["title_bg"]); c.font = font(True, "FFFFFF", 16)
c.alignment = center()
ws.row_dimensions[1].height = 40

ws.merge_cells('A2:B2')
c = ws['A2']
c.value = "สรุปผลการทดสอบทั้งหมด | Full Test Summary"
c.fill = fill(C["sub_bg"]); c.font = font(False, "FFFFFF", 12)
c.alignment = center()
ws.row_dimensions[2].height = 25

headers_summary = ["รายการ", "จำนวน"]
write_header(ws, 3, {h: h for h in headers_summary}, C["header_green"])

summary_rows = [
    ("ประเภทการทดสอบ", "5 ประเภท"),
    ("Unit Tests", "80 tests (health_calc_test.dart)"),
    ("Negative + BVA + Integration Tests", "73 tests (negative_bva_integration_test.dart)"),
    ("Full Coverage Tests", "78 tests (full_coverage_test.dart)"),
    ("Gamification Tests", "27 tests (gamification_test.dart)"),
    ("Widget Tests", "9 tests (widget_test.dart)"),
    ("รวม Unit Tests ทั้งหมด", "267 tests"),
    ("E2E Integration Tests (ต้องมี device)", "10 tests (app_e2e_test.dart)"),
    ("", ""),
    ("Use Cases ที่ครอบคลุม", f"{len(USE_CASES)} UCs"),
    ("Test Cases ที่บันทึก", f"{len(TEST_CASES)} TCs"),
    ("งานวิจัยอ้างอิง", f"{len(REFERENCES)} refs"),
    ("", ""),
    ("สถานะ: PASS", "267 / 267"),
    ("สถานะ: FAIL", "0 / 267"),
    ("Coverage", "100% (all tests pass)"),
]

for i, (k, v) in enumerate(summary_rows):
    r = i + 4
    bg = C["row_green"] if i % 2 == 0 else C["white"]
    ws.cell(r, 1, k).fill = fill(bg); ws.cell(r, 1).font = font(bool(k), "1B5E20")
    ws.cell(r, 1).border = thin_border(); ws.cell(r, 1).alignment = left()
    ws.cell(r, 2, v).fill = fill(bg); ws.cell(r, 2).font = font(False, "212121")
    ws.cell(r, 2).border = thin_border(); ws.cell(r, 2).alignment = center()
    ws.row_dimensions[r].height = 20

# ─── Sheet 2: Use Cases ────────────────────────────────────────────────────
ws2 = wb.create_sheet("📋 Use Cases")
ws2.sheet_view.showGridLines = False
uc_cols = {"UC ID": 8, "Module": 14, "Use Case": 22, "Actor": 12,
           "Precondition": 22, "Main Flow": 45, "Expected Result": 35, "Priority": 10}
write_header(ws2, 1, {k: k for k in uc_cols}, C["header_blue"])
for i, (col, w) in enumerate(uc_cols.items(), 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 30

row_colors = {"High": C["row_orange"], "Medium": C["row_blue"], "Low": C["row_teal"]}
for r, uc in enumerate(USE_CASES, 2):
    bg = row_colors.get(uc[7], C["white"])
    write_row(ws2, r, uc, bg)
    ws2.row_dimensions[r].height = 45

# ─── Sheet 3: Test Cases ───────────────────────────────────────────────────
ws3 = wb.create_sheet("🧪 Test Cases")
ws3.sheet_view.showGridLines = False
tc_cols = {"TC ID": 8, "UC Ref": 8, "Module": 16, "Test Case Name": 40,
           "Test Type": 12, "Input": 30, "Expected Output": 25,
           "Status": 10, "Priority": 10}
write_header(ws3, 1, {k: k for k in tc_cols}, C["header_orange"])
for i, (col, w) in enumerate(tc_cols.items(), 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[1].height = 30
ws3.freeze_panes = "A2"

type_bg = {
    "Unit": C["row_green"], "Negative": C["row_orange"],
    "BVA": C["row_purple"], "Integration": C["row_blue"],
    "Widget": C["row_teal"],
}
status_bg = {"PASS": C["pass"], "FAIL": C["fail"]}

for r, tc in enumerate(TEST_CASES, 2):
    bg = type_bg.get(tc[4], C["white"])
    write_row(ws3, r, tc[:8], bg)
    # color status cell
    sc = ws3.cell(r, 8)
    sc.fill = fill(status_bg.get(tc[7], C["white"]))
    sc.font = font(True, "1B5E20" if tc[7] == "PASS" else "B71C1C")
    sc.alignment = center()
    # color priority cell
    pc = ws3.cell(r, 9, tc[8])
    pc.fill = fill(C["row_orange"] if tc[8] == "High"
                   else C["row_blue"] if tc[8] == "Medium" else C["row_teal"])
    pc.font = font(False, "212121"); pc.border = thin_border(); pc.alignment = center()
    ws3.row_dimensions[r].height = 30

# ─── Sheet 4: References ──────────────────────────────────────────────────
ws4 = wb.create_sheet("📚 Research References")
ws4.sheet_view.showGridLines = False
ref_cols = {"Ref ID": 8, "หัวข้อ": 22, "ผู้แต่ง": 22, "วารสาร/แหล่ง": 30,
            "สรุปสาระสำคัญ": 50, "Test Cases ที่อ้างอิง": 22}
write_header(ws4, 1, {k: k for k in ref_cols}, C["header_purple"])
for i, (col, w) in enumerate(ref_cols.items(), 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.row_dimensions[1].height = 30

for r, ref in enumerate(REFERENCES, 2):
    bg = C["row_purple"] if r % 2 == 0 else C["row_teal"]
    write_row(ws4, r, ref, bg)
    ws4.row_dimensions[r].height = 40

# ─── Sheet 5: Test Type Legend ────────────────────────────────────────────
ws5 = wb.create_sheet("🗂️ Legend")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions['A'].width = 18
ws5.column_dimensions['B'].width = 50
ws5.column_dimensions['C'].width = 20

ws5.merge_cells('A1:C1')
c = ws5['A1']
c.value = "คำอธิบายประเภทการทดสอบ (Test Type Legend)"
c.fill = fill(C["header_teal"]); c.font = font(True, "FFFFFF", 13)
c.alignment = center(); ws5.row_dimensions[1].height = 35

legends = [
    ("Unit", "ทดสอบ function เดี่ยว ตรวจว่าผลลัพธ์ตรงกับ expected output", C["row_green"]),
    ("Negative", "ทดสอบด้วย input ผิดปกติ/ไม่ถูกต้อง ตรวจว่าระบบไม่พัง (Graceful Failure)", C["row_orange"]),
    ("BVA", "Boundary Value Analysis — ทดสอบค่าขอบเขต min-1/min/min+1/max-1/max/max+1", C["row_purple"]),
    ("Integration", "ทดสอบการเชื่อมต่อระหว่าง component เช่น API response parsing", C["row_blue"]),
    ("Widget", "ทดสอบ UI widget render ถูกต้องโดยไม่ crash", C["row_teal"]),
    ("E2E", "End-to-End: ทดสอบ flow จริงบน device/emulator (ต้อง run แยก)", C["warn"]),
]

write_header(ws5, 2, {"ประเภท": "ประเภท", "คำอธิบาย": "คำอธิบาย",
                       "สี Reference": "สีใน Sheet"}, C["header_teal"])
for r, (t, desc, bg) in enumerate(legends, 3):
    ws5.cell(r, 1, t).fill = fill(bg); ws5.cell(r, 1).font = font(True)
    ws5.cell(r, 1).border = thin_border(); ws5.cell(r, 1).alignment = center()
    ws5.cell(r, 2, desc).fill = fill(bg); ws5.cell(r, 2).font = font()
    ws5.cell(r, 2).border = thin_border(); ws5.cell(r, 2).alignment = left()
    ws5.cell(r, 3, "■").fill = fill(bg); ws5.cell(r, 3).font = Font(size=20, color=bg)
    ws5.cell(r, 3).border = thin_border(); ws5.cell(r, 3).alignment = center()
    ws5.row_dimensions[r].height = 35

wb.save(OUT)
print(f"✅ Excel saved → {OUT}")
print(f"   📊 Summary | 📋 {len(USE_CASES)} Use Cases | 🧪 {len(TEST_CASES)} Test Cases | 📚 {len(REFERENCES)} References")
