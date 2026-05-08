#!/usr/bin/env python3
"""
Generate backend/migrations/v28_thaifcd_import_foods.sql from:
  - Downloads/ingredient.txt (INSERT statements, food_type 'ingredient' -> raw_ingredient)
  - Downloads/beverage.xlsx (ThaiFCD sheet)

Run from repo root (paths below are configurable):
  python backend/scripts/generate_v28_thaifcd_migration.py

Requires: openpyxl
"""

from __future__ import annotations

import os
from pathlib import Path
import re

try:
    import openpyxl
except ImportError as e:
    raise SystemExit("pip install openpyxl") from e

REPO = Path(__file__).resolve().parents[2]

# Prefer checked-in seeds; override via env for local/regenerate from Downloads.
_ING_DEFAULT = REPO / "backend" / "seeds" / "thaifcd" / "ingredient.txt"
_BEV_DEFAULT = REPO / "backend" / "seeds" / "thaifcd" / "beverage.xlsx"
ING_TXT = Path(os.getenv("CG_IMPORT_INGREDIENT", str(_ING_DEFAULT)))
BEV_XLSX = Path(os.getenv("CG_IMPORT_BEVERAGE", str(_BEV_DEFAULT)))
OUT_SQL = REPO / "backend" / "migrations" / "v28_thaifcd_import_foods.sql"


def parse_ingredient_line(line: str) -> dict | None:
    line = line.strip()
    if not line.startswith("INSERT"):
        return None
    # Support "VALUES (...)" or "VALUES(...)"
    m = re.search(r"\bVALUES\s*\(", line, flags=re.I)
    if not m:
        return None
    i = m.end()  # immediately after VALUES' opening '('
    depth = 1
    quote = False
    start = i
    end_paren = None
    k = i
    while k < len(line):
        ch = line[k]
        if ch == "'":
            if quote and k + 1 < len(line) and line[k + 1] == "'":
                k += 2
                continue
            quote = not quote
            k += 1
            continue
        if not quote:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
                if depth == 0:
                    end_paren = k
                    break
        k += 1
    if end_paren is None:
        return None
    tup = line[start:end_paren]

    vals: list[str] = []
    cur: list[str] = []
    quote = False
    k = 0
    while k < len(tup):
        ch = tup[k]
        if ch == "'":
            if quote and k + 1 < len(tup) and tup[k + 1] == "'":
                cur.append("'")
                k += 2
                continue
            quote = not quote
            cur.append(ch)
            k += 1
            continue
        if ch == "," and not quote:
            vals.append("".join(cur).strip())
            cur = []
            k += 1
            continue
        cur.append(ch)
        k += 1
    vals.append("".join(cur).strip())
    if len(vals) != 11:
        return None

    def parse_atom(s: str):
        s = s.strip()
        if s.startswith("'") and s.endswith("'"):
            return s[1:-1].replace("''", "'")
        if s.upper() == "NULL":
            return None
        try:
            return float(s)
        except ValueError:
            return s

    return {
        "food_name": parse_atom(vals[0]),
        "food_type_raw": parse_atom(vals[1]),
        "calories": parse_atom(vals[2]),
        "protein": parse_atom(vals[3]),
        "carbs": parse_atom(vals[4]),
        "fat": parse_atom(vals[5]),
        "sodium": parse_atom(vals[6]),
        "sugar": parse_atom(vals[7]),
        "cholesterol": parse_atom(vals[8]),
        "fiber_g": parse_atom(vals[9]),
        "serving_quantity": parse_atom(vals[10]),
    }


def sql_str(s: str) -> str:
    return "'" + str(s).replace("'", "''") + "'"


def sql_num(x) -> str:
    if x is None:
        return "NULL"
    if isinstance(x, float) and x == int(x):
        return str(int(x))
    return str(x)


def load_ingredients() -> list[dict]:
    txt = ING_TXT.read_text(encoding="utf-8")
    rows = []
    for line in txt.splitlines():
        p = parse_ingredient_line(line)
        if p:
            rows.append(p)
    # First occurrence wins (dedupe duplicate food_name in source file).
    by_name: dict = {}
    for r in rows:
        if r["food_name"] not in by_name:
            by_name[r["food_name"]] = r
    return list(by_name.values())


def fnum(cell) -> float | None:
    if cell is None or cell == "" or cell == "-":
        return None
    try:
        return float(str(cell).strip())
    except ValueError:
        return None


def load_beverages() -> list[dict]:
    wb = openpyxl.load_workbook(str(BEV_XLSX), read_only=True)
    sh = wb[wb.sheetnames[0]]
    out = []
    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        code = row[0]
        name = row[1]
        if not name:
            continue
        prot = fnum(row[2]) or 0.0
        fat = fnum(row[3]) or 0.0
        carb = fnum(row[4])
        ener = fnum(row[5])
        total_e = fnum(row[6])
        sug = fnum(row[7])
        kcal = ener if ener is not None else total_e
        if kcal is None:
            continue
        if carb is None:
            carb = 0.0
        out.append(
            {
                "food_code": str(code).strip(),
                "food_name": str(name).strip(),
                "protein": prot,
                "fat": fat,
                "carbs": carb,
                "calories": kcal,
                "sugar": sug,
                "fiber_g": 0.0,
            }
        )
    deduped: dict = {}
    for b in out:
        if b["food_name"] not in deduped:
            deduped[b["food_name"]] = b
    return list(deduped.values())


HEADER = """-- v28: Import / upsert ThaiFCD-derived foods from local seed files.
--
-- Sources:
--   * ingredient.txt  -> cleangoal.foods (canonical_food_type raw_ingredient)
--   * beverage.xlsx   -> cleangoal.foods (beverage) + cleangoal.beverages row
--
-- Generated by: backend/scripts/generate_v28_thaifcd_migration.py
-- Idempotent via ON CONFLICT (food_name) DO UPDATE.

BEGIN;

SET search_path TO cleangoal, public;

DO $ensure_food_name_uq$
BEGIN
  IF NOT EXISTS (
    SELECT 1 FROM pg_constraint
    WHERE conname = 'uq_foods_food_name'
      AND conrelid = 'cleangoal.foods'::regclass
  ) THEN
    ALTER TABLE cleangoal.foods
      ADD CONSTRAINT uq_foods_food_name UNIQUE (food_name);
  END IF;
END
$ensure_food_name_uq$;

-- Dedicated taxonomy buckets (avoid clashing ad-hoc category rows).
INSERT INTO cleangoal.dish_categories (category_name, canonical_food_type, display_order, description)
VALUES
    ('ThaiFCD import — วัตถุดิบ', 'raw_ingredient'::cleangoal.food_type, 990, 'Imported from ingredient.txt seed'),
    ('ThaiFCD import — เครื่องดื่ม', 'beverage'::cleangoal.food_type, 991, 'Imported from beverage.xlsx (per 100 ml)')
ON CONFLICT (category_name, canonical_food_type) DO NOTHING;

-- ─── Ingredients: dishes + foods ───────────────────────────────────────────
"""

DISH_ING_TEMPLATE = """
INSERT INTO cleangoal.dishes (dish_name, dish_category_id, canonical_food_type, cuisine, description)
SELECT {fname},
       dc.dish_category_id,
       'raw_ingredient'::cleangoal.food_type,
       'Thai',
       'ThaiFCD ingredient seed'
FROM cleangoal.dish_categories dc
WHERE dc.category_name = 'ThaiFCD import — วัตถุดิบ'
  AND dc.canonical_food_type = 'raw_ingredient'::cleangoal.food_type
ON CONFLICT (dish_name, dish_category_id) DO NOTHING;
"""

FOOD_ING_BLOCK = """
INSERT INTO cleangoal.foods (
    food_name, food_type, calories, protein, carbs, fat, sodium, sugar, cholesterol, fiber_g,
    serving_quantity, serving_unit_id, dish_id,
    updated_at
)
SELECT
    v.food_name,
    'raw_ingredient'::cleangoal.food_type,
    v.calories, v.protein, v.carbs, v.fat, v.sodium, v.sugar, v.cholesterol,
    COALESCE(v.fiber_g, 0),
    v.serving_quantity,
    (SELECT unit_id FROM cleangoal.units WHERE lower(name) = 'g' LIMIT 1),
    d.dish_id,
    NOW()
FROM (
    VALUES
__ROWS__
) AS v(food_name, calories, protein, carbs, fat, sodium, sugar, cholesterol, fiber_g, serving_quantity)
JOIN cleangoal.dishes d
  ON d.dish_name = v.food_name
JOIN cleangoal.dish_categories dc
  ON dc.dish_category_id = d.dish_category_id
 AND dc.category_name = 'ThaiFCD import — วัตถุดิบ'
ON CONFLICT (food_name) DO UPDATE SET
    food_type    = EXCLUDED.food_type,
    calories     = EXCLUDED.calories,
    protein      = EXCLUDED.protein,
    carbs        = EXCLUDED.carbs,
    fat          = EXCLUDED.fat,
    sodium       = EXCLUDED.sodium,
    sugar        = EXCLUDED.sugar,
    cholesterol  = EXCLUDED.cholesterol,
    fiber_g      = EXCLUDED.fiber_g,
    serving_quantity = EXCLUDED.serving_quantity,
    serving_unit_id = EXCLUDED.serving_unit_id,
    dish_id      = EXCLUDED.dish_id,
    deleted_at   = NULL,
    updated_at   = NOW();
"""

DISH_BEV_ONE = """
INSERT INTO cleangoal.dishes (dish_name, dish_category_id, canonical_food_type, cuisine, description)
SELECT {fname},
       dc.dish_category_id,
       'beverage'::cleangoal.food_type,
       'Thai',
       {fdesc}
FROM cleangoal.dish_categories dc
WHERE dc.category_name = 'ThaiFCD import — เครื่องดื่ม'
  AND dc.canonical_food_type = 'beverage'::cleangoal.food_type
ON CONFLICT (dish_name, dish_category_id) DO NOTHING;
"""

FOOD_BEV_BLOCK = """
INSERT INTO cleangoal.foods (
    food_name, food_type, calories, protein, carbs, fat, sodium, sugar, cholesterol, fiber_g,
    serving_quantity, serving_unit_id, dish_id,
    updated_at
)
SELECT
    v.food_name,
    'beverage'::cleangoal.food_type,
    v.calories, v.protein, v.carbs, v.fat,
    NULL,
    v.sugar,
    NULL,
    COALESCE(v.fiber_g, 0),
    100,
    (SELECT unit_id FROM cleangoal.units WHERE lower(name) = 'ml' LIMIT 1),
    d.dish_id,
    NOW()
FROM (
    VALUES
__ROWS__
) AS v(food_name, food_code, calories, protein, carbs, fat, sugar, fiber_g)
JOIN cleangoal.dishes d ON d.dish_name = v.food_name
JOIN cleangoal.dish_categories dc
  ON dc.dish_category_id = d.dish_category_id
 AND dc.category_name = 'ThaiFCD import — เครื่องดื่ม'
ON CONFLICT (food_name) DO UPDATE SET
    food_type    = EXCLUDED.food_type,
    calories     = EXCLUDED.calories,
    protein      = EXCLUDED.protein,
    carbs        = EXCLUDED.carbs,
    fat          = EXCLUDED.fat,
    sugar        = EXCLUDED.sugar,
    fiber_g      = EXCLUDED.fiber_g,
    serving_quantity = EXCLUDED.serving_quantity,
    serving_unit_id = EXCLUDED.serving_unit_id,
    dish_id      = EXCLUDED.dish_id,
    deleted_at   = NULL,
    updated_at   = NOW();
"""

BEV_EXTRA = """
-- Beverage dimension (water tracking): 100 ml per catalogue row nutrition basis.
INSERT INTO cleangoal.beverages (
    food_id, volume_ml, is_alcoholic, caffeine_mg, sugar_level_label, container_type
)
SELECT
    f.food_id,
    100,
    FALSE,
    0,
    CASE
        WHEN f.sugar IS NULL THEN 'ไม่ระบุ'
        WHEN f.sugar >= 12 THEN 'หวาน'
        WHEN f.sugar >= 6  THEN 'หวานปานกลาง'
        ELSE 'หวานน้อย'
    END,
    CASE WHEN f.food_name ILIKE '%กระป๋อง%' OR f.food_name ILIKE '%กระปอก%' THEN 'กระป๋อง' ELSE 'แก้ว' END
FROM cleangoal.foods f
JOIN cleangoal.dishes d ON d.dish_id = f.dish_id
JOIN cleangoal.dish_categories dc ON dc.dish_category_id = d.dish_category_id
WHERE dc.category_name = 'ThaiFCD import — เครื่องดื่ม'
  AND f.deleted_at IS NULL
  AND NOT EXISTS (
      SELECT 1 FROM cleangoal.beverages b WHERE b.food_id = f.food_id
  );
"""

FOOTER = """
INSERT INTO cleangoal.schema_migrations(version)
VALUES ('v28_thaifcd_import_foods')
ON CONFLICT (version) DO NOTHING;

COMMIT;
"""


def chunk_values(rows_sql: list[str], max_lines: int = 80) -> list[str]:
    parts = []
    for i in range(0, len(rows_sql), max_lines):
        parts.append(",\n".join(rows_sql[i : i + max_lines]))
    return parts


def main():
    ingredients = sorted(load_ingredients(), key=lambda r: r["food_name"])
    beverages = sorted(load_beverages(), key=lambda r: r["food_name"])

    lines = [HEADER]
    lines.append("-- Dishes per ingredient food_name")
    for r in ingredients:
        lines.append(DISH_ING_TEMPLATE.format(fname=sql_str(r["food_name"])))

    ing_val_lines = []
    for r in ingredients:
        ing_val_lines.append(
            "("
            + ", ".join(
                [
                    sql_str(r["food_name"]),
                    sql_num(r["calories"]),
                    sql_num(r["protein"]),
                    sql_num(r["carbs"]),
                    sql_num(r["fat"]),
                    sql_num(r["sodium"]),
                    sql_num(r["sugar"]),
                    sql_num(r["cholesterol"]),
                    sql_num(r["fiber_g"]),
                    sql_num(r["serving_quantity"]),
                ]
            )
            + ")"
        )
    blocks = chunk_values(ing_val_lines, 80)
    for idx, blk in enumerate(blocks):
        label = FOOD_ING_BLOCK.replace("__ROWS__", blk)
        lines.append(f"-- Ingredients foods batch {idx+1}/{len(blocks)}")
        lines.append(label)

    lines.append("-- Beverage dishes")
    for b in beverages:
        lines.append(
            DISH_BEV_ONE.format(
                fname=sql_str(b["food_name"]),
                fdesc=sql_str(f"ThaiFCD {b['food_code']}"),
            )
        )

    bev_val_lines = []
    for b in beverages:
        bev_val_lines.append(
            "("
            + ", ".join(
                [
                    sql_str(b["food_name"]),
                    sql_str(b["food_code"]),
                    sql_num(b["calories"]),
                    sql_num(b["protein"]),
                    sql_num(b["carbs"]),
                    sql_num(b["fat"]),
                    sql_num(b["sugar"]),
                    sql_num(b["fiber_g"]),
                ]
            )
            + ")"
        )
    bb = chunk_values(bev_val_lines, 80)
    for idx, blk in enumerate(bb):
        lines.append(f"-- Beverage foods batch {idx+1}/{len(bb)}")
        lines.append(FOOD_BEV_BLOCK.replace("__ROWS__", blk))

    lines.append(BEV_EXTRA)
    lines.append(FOOTER)

    OUT_SQL.parent.mkdir(parents=True, exist_ok=True)
    OUT_SQL.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT_SQL} (ingredients {len(ingredients)}, beverages {len(beverages)})")


if __name__ == "__main__":
    main()
